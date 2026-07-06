"""Fill the review spreadsheet (Sheet1) from a local, offline document review.

Bridges the folder-review engine with the "Option A2 ..." workbook:

* Each immediate subfolder under ``--folders-root`` is one person.
* Every folder is matched to a Sheet1 row using, in priority order:
    1. the passport number read from the front page   -> column E
    2. the TSF ID embedded in a file name             -> column B
    3. the Booking ID embedded in a file name         -> column A
    4. the person name                                -> column D  (last resort)
* For a matched row the review results are written into the row:
    - Passport No. Back            -> address/last-page review result
    - Passport No. Blank Pages     -> blank-pages review result
    - Photo Dimension (390 x 567)  -> whole-photo review result
    - Any comments for errors      -> consolidated comments (masked numbers only)
  using the words ``ok`` / ``Error`` / ``Not legible`` from the process document.
* The matched folder is renamed to the row's Folder Name (column L) and its files
  to the row's file names (columns M..P), in place, and every rename is logged to
  ``renames.csv`` so it can be undone.

The original workbook is never modified: results are written to a new copy.
Full passport numbers and names are never printed to the console unless
``--show-pii-in-console`` is passed, and never leave the machine.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from .config import ReviewConfig
from .documents import normalize_name
from .dynamic_discovery import list_person_folders
from .local_vision import create_visual_analyzer
from .models import ManifestRow, PersonReview, Status
from .network_guard import enable_strict_offline_mode
from .ocr import MemoizingOCREngine, create_ocr_engine, normalize_alphanumeric
from .pipeline import _type_from_filename, export_bundle, review_person_folder
from .reporting import append_result_row, write_manual_queue, write_results, write_summary


_TSF_RE = re.compile(r"TSF\d{4,}", re.IGNORECASE)
_BOOKING_RE = re.compile(r"\d{4}-[A-Za-z]{2,6}-\d{3,5}")


def na(value: object) -> str:
    return normalize_alphanumeric(str(value or ""))


# ---------------------------------------------------------------------------
# Spreadsheet model
# ---------------------------------------------------------------------------

@dataclass(slots=True)
class SheetRow:
    row_number: int
    name: str = ""
    booking_id: str = ""
    tsf_id: str = ""
    passport_front: str = ""
    folder_name: str = ""
    file_front: str = ""
    file_last: str = ""
    file_blank: str = ""
    file_photo: str = ""
    filled: bool = False


@dataclass(slots=True)
class SheetColumns:
    passport_back: int
    passport_blank: int
    photo_dimension: int
    comments: int
    booking_id: int | None = None
    tsf_id: int | None = None
    name: int | None = None
    passport_front: int | None = None
    folder_name: int | None = None
    file_front: int | None = None
    file_last: int | None = None
    file_blank: int | None = None
    file_photo: int | None = None


# Fallback 1-based column positions matching the shipped template, used only when a
# header cannot be located by text.
_DEFAULT_COLS = {
    "booking_id": 1,
    "tsf_id": 2,
    "name": 4,
    "passport_front": 5,
    "passport_back": 6,
    "passport_blank": 7,
    "photo_dimension": 8,
    "comments": 9,
    "folder_name": 12,
    "file_front": 13,
    "file_last": 14,
    "file_blank": 15,
    "file_photo": 16,
}


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _find_header_row(ws, max_scan: int = 8) -> int:
    for row in range(1, min(max_scan, ws.max_row or max_scan) + 1):
        texts = {_cell_text(ws.cell(row=row, column=c).value).lower() for c in range(1, (ws.max_column or 0) + 1)}
        if any("passport no" in t for t in texts):
            return row
    return 1


def _resolve_columns(ws, header_row: int) -> SheetColumns:
    headers: dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        text = _cell_text(ws.cell(row=header_row, column=c).value).lower()
        if text:
            headers.setdefault(text, c)

    def by_exact(*names: str) -> int | None:
        for name in names:
            if name in headers:
                return headers[name]
        return None

    def by_contains(*needles: str) -> int | None:
        for text, col in headers.items():
            if all(n in text for n in needles):
                return col
        return None

    def resolve(key: str, exact: tuple[str, ...] = (), contains: tuple[str, ...] = ()) -> int | None:
        found = by_exact(*exact)
        if found is None and contains:
            found = by_contains(*contains)
        return found if found is not None else _DEFAULT_COLS.get(key)

    return SheetColumns(
        booking_id=resolve("booking_id", exact=("booking id",), contains=("booking",)),
        tsf_id=resolve("tsf_id", exact=("tsf id",), contains=("tsf",)),
        name=resolve("name", exact=("name",)),
        passport_front=resolve("passport_front", exact=("passport no. front",), contains=("passport no", "front")),
        passport_back=resolve("passport_back", exact=("passport no. back",), contains=("passport no", "back")) or _DEFAULT_COLS["passport_back"],
        passport_blank=resolve("passport_blank", exact=("passport no. blank pages",), contains=("passport no", "blank")) or _DEFAULT_COLS["passport_blank"],
        photo_dimension=resolve("photo_dimension", contains=("photo dimension",)) or _DEFAULT_COLS["photo_dimension"],
        comments=resolve("comments", contains=("comments",)) or _DEFAULT_COLS["comments"],
        folder_name=resolve("folder_name", exact=("folder name",)),
        file_front=resolve("file_front", contains=("file name", "front")),
        file_last=resolve("file_last", contains=("file name", "last")),
        file_blank=resolve("file_blank", contains=("file name", "blank")),
        file_photo=resolve("file_photo", contains=("file name", "photo")),
    )


def _front_review_column(ws, header_row: int) -> int:
    """Locate the 'Passport Front Page Review' column, or append it after the last
    column (appending avoids shifting the =CONCATENATE(...) formula columns)."""
    targets = {"passport front page review", "passport no. front review", "front page review", "passport front page"}
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        if _cell_text(ws.cell(row=header_row, column=c).value).lower() in targets:
            return c
    col = max_col + 1
    ws.cell(row=header_row, column=col).value = "Passport Front Page Review"
    return col


def _email_column(ws, header_row: int) -> int:
    """Locate the 'Email' column, or append it after the last column."""
    targets = {"email", "email section", "re-upload email", "re upload email", "email (re-upload request)"}
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        if _cell_text(ws.cell(row=header_row, column=c).value).lower() in targets:
            return c
    col = max_col + 1
    ws.cell(row=header_row, column=col).value = "Email (re-upload request)"
    return col


def _load_sheet_rows(ws, header_row: int, cols: SheetColumns) -> list[SheetRow]:
    rows: list[SheetRow] = []
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        def get(col: int | None) -> str:
            return _cell_text(ws.cell(row=r, column=col).value) if col else ""

        row = SheetRow(
            row_number=r,
            name=get(cols.name),
            booking_id=get(cols.booking_id),
            tsf_id=get(cols.tsf_id),
            passport_front=get(cols.passport_front),
            folder_name=get(cols.folder_name),
            file_front=get(cols.file_front),
            file_last=get(cols.file_last),
            file_blank=get(cols.file_blank),
            file_photo=get(cols.file_photo),
        )
        if any((row.name, row.passport_front, row.folder_name, row.tsf_id, row.booking_id)):
            rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

@dataclass(slots=True)
class MatchIndex:
    by_number: dict[str, list[SheetRow]] = field(default_factory=dict)
    by_tsf: dict[str, list[SheetRow]] = field(default_factory=dict)
    by_booking: dict[str, list[SheetRow]] = field(default_factory=dict)
    by_name: dict[str, list[SheetRow]] = field(default_factory=dict)

    @classmethod
    def build(cls, rows: list[SheetRow]) -> "MatchIndex":
        index = cls()
        for row in rows:
            if row.passport_front:
                index.by_number.setdefault(na(row.passport_front), []).append(row)
            if row.tsf_id:
                index.by_tsf.setdefault(na(row.tsf_id), []).append(row)
            if row.booking_id:
                index.by_booking.setdefault(na(row.booking_id), []).append(row)
            if row.name:
                index.by_name.setdefault(normalize_name(row.name), []).append(row)
        return index


def _single(rows: list[SheetRow] | None) -> SheetRow | None:
    """Return a single unfilled row, or None when absent or ambiguous."""
    if not rows:
        return None
    unfilled = [row for row in rows if not row.filled]
    if len(unfilled) == 1:
        return unfilled[0]
    if not unfilled and len(rows) == 1:
        return rows[0]  # already filled; let the caller flag the duplicate
    return None


def _booking_row(index: MatchIndex, booking_ids: set[str]) -> SheetRow | None:
    """Booking IDs on file names (e.g. 2608-KMYM-0050) may lack the workbook's
    leading '#'/trailing '-N', so match on a shared prefix."""
    for booking in booking_ids:
        for key, rows in index.by_booking.items():
            row = _single(rows)
            if row is None:
                continue
            if key == booking or key.startswith(booking) or booking.startswith(key):
                return row
    return None


def _match_by_name(folder: Path, rows: list[SheetRow]) -> SheetRow | None:
    normalized_folder = normalize_name(folder.name)
    if not normalized_folder:
        return None
    tokens = set(normalized_folder.split())
    candidates: list[SheetRow] = []
    for row in rows:
        if row.filled or not row.name:
            continue
        name_tokens = set(normalize_name(row.name).split())
        if not name_tokens:
            continue
        # The folder name is a short form; require the folder tokens to be a subset
        # of the row's name tokens (or vice versa) so a shared first name alone is
        # not treated as a match.
        if tokens <= name_tokens or name_tokens <= tokens:
            candidates.append(row)
    if len(candidates) == 1:
        return candidates[0]
    return None


def _ids_from_folder(folder: Path) -> tuple[set[str], set[str]]:
    tsf_ids: set[str] = set()
    booking_ids: set[str] = set()
    for path in folder.rglob("*"):
        if not path.is_file():
            continue
        for match in _TSF_RE.findall(path.name):
            tsf_ids.add(na(match))
        for match in _BOOKING_RE.findall(path.name):
            booking_ids.add(na(match))
    return tsf_ids, booking_ids


@dataclass(slots=True)
class Decision:
    row: SheetRow | None
    method: str = ""
    note: str = ""
    confident: bool = False


def _decide_row(
    detected_number: str,
    tentative: SheetRow | None,
    tsf_ids: set[str],
    booking_ids: set[str],
    folder: Path,
    index: MatchIndex,
    rows: list[SheetRow],
) -> Decision:
    number_row = _single(index.by_number.get(na(detected_number))) if detected_number else None
    tsf_row = _single([r for tsf in tsf_ids for r in index.by_tsf.get(tsf, [])]) if tsf_ids else None
    notes: list[str] = []

    # Priority: passport number (authoritative) -> TSF -> Booking -> name.
    if number_row is not None:
        chosen, method, confident = number_row, "passport number", True
        if tsf_row is not None and tsf_row is not number_row:
            notes.append("The TSF ID on the files points to a different row than the passport number; used the passport number.")
    elif tsf_row is not None:
        chosen, method, confident = tsf_row, "TSF ID", True
    elif tentative is not None:
        chosen, method, confident = tentative, "TSF/Booking ID", True
    else:
        booking_row = _booking_row(index, booking_ids)
        if booking_row is not None:
            chosen, method, confident = booking_row, "booking ID", True
        else:
            name_row = _match_by_name(folder, rows)
            if name_row is not None:
                chosen, method, confident = name_row, "person name", False
                notes.append("Matched by person name only; confirm identity before relying on the rename.")
            else:
                return Decision(row=None)

    if chosen.filled:
        notes.append("This row was already filled by another folder; possible duplicate submission.")
    if detected_number and chosen.passport_front and na(detected_number) != na(chosen.passport_front):
        notes.append("The passport number read from the images does not match this row's number on record; verify.")

    return Decision(row=chosen, method=method, note=" ".join(notes), confident=confident)


# ---------------------------------------------------------------------------
# Result mapping
# ---------------------------------------------------------------------------

def _cell_word(status: Status) -> str:
    """Map an internal status onto the process document's three outcome words."""
    if status in (Status.OK, Status.NOT_APPLICABLE):
        return "ok"
    if status == Status.NOT_LEGIBLE:
        return "Not legible"
    # Error, Missing and any residual manual-review doubt route to the human second
    # check as an error with an explanatory comment.
    return "Error"


def _document_cell(status: Status, sources: list) -> str:
    """Cell value for one document column. A PNG source is an ineligible format, so
    it is reported as 'Error' (the reason is spelled out in the comments). A
    document represented only by a PDF/other file is already Missing -> 'Error'."""
    if any(source.path.suffix.lower() == ".png" for source in (sources or [])):
        return "Error"
    return _cell_word(status)


_RECORD_COMMENT_PREFIXES = (
    "Detected passport number",
    "The passport number on record",
    "Unclassified files:",
    "Could not copy",
    "Format:",
)


def _sheet_comments(review: PersonReview, match_note: str) -> str:
    parts: list[str] = []
    if match_note:
        parts.append(match_note)
    for comment in review.comments:
        if comment.startswith(_RECORD_COMMENT_PREFIXES):
            parts.append(comment)
    for label, result, always in (
        ("Front", review.front, False),
        ("Back", review.back, True),
        ("Blank pages", review.blank, True),
        ("Photo", review.photo, True),
    ):
        if result.comments and (always or result.status != Status.OK):
            parts.append(f"{label}: " + "; ".join(result.comments))
    seen: set[str] = set()
    ordered: list[str] = []
    for part in parts:
        if part and part not in seen:
            seen.add(part)
            ordered.append(part)
    return " | ".join(ordered)


def _has_quality_issue(result) -> bool:
    """True if the document was flagged for image quality (blur / low contrast),
    which a re-upload could fix - as opposed to only an unreadable number."""
    return any(("blur" in comment.lower() or "contrast" in comment.lower()) for comment in result.comments)


def _email_reason(key: str, cell: str, png: bool, ineligible: bool, missing: bool) -> str:
    """Applicant-friendly reason for re-uploading one document."""
    if png:
        return "the file is a PNG image; please re-upload it as a JPG/JPEG file."
    if ineligible:
        return "the file is a PDF (or another format); please re-upload it as a JPG/JPEG image."
    if missing:
        return "the document is missing; please upload it as a JPG/JPEG image."
    if cell == "Not legible":
        return "the image is not clear/legible; please re-upload a clearer JPG/JPEG scan."
    if key == "photo":
        return ("the photo does not meet the requirement (must be 390 x 567 px JPG at 300 DPI, plain white "
                "background, face clearly visible); please re-upload a compliant passport photo.")
    return "there is an issue with this document; please re-upload a clear JPG/JPEG file."


def _email_text(name: str, review: PersonReview, config: ReviewConfig) -> str:
    """Compose a re-upload request listing only the documents that need attention.
    Blank pages that merely could not be OCR-verified are left for manual review."""
    bundle = review.bundle
    ineligible_types = {(_type_from_filename(p, config) or "document") for p in (bundle.ineligible if bundle else [])}
    lines: list[str] = []
    for label, key, result, sources in (
        ("Front page", "front", review.front, bundle.front if bundle else []),
        ("Back page (address)", "back", review.back, bundle.back if bundle else []),
        ("Blank pages", "blank", review.blank, bundle.blank if bundle else []),
        ("Photo", "photo", review.photo, bundle.photo if bundle else []),
    ):
        cell = _document_cell(result.status, sources)
        if cell == "ok":
            continue
        # "Not legible" with no image-quality problem means the number simply could
        # not be read (dot-printed/mirrored blank-page number, or the small
        # below-barcode number). That is a manual-verify item, not a re-upload.
        if cell == "Not legible" and not _has_quality_issue(result):
            continue
        png = any(s.path.suffix.lower() == ".png" for s in sources)
        missing = not sources
        ineligible = missing and key in ineligible_types
        # Blank pages are only put in the email when missing or in a wrong format -
        # never for image quality (they are verified manually).
        if key == "blank" and not (png or ineligible or missing):
            continue
        lines.append(f"- {label}: {_email_reason(key, cell, png, ineligible, missing)}")

    if not lines:
        return ""
    greeting = f"Dear {name}," if name else "Dear Applicant,"
    return greeting + "\n\nPlease re-upload the following document(s):\n" + "\n".join(lines) + "\n\nThank you."


def _file_actions(review: PersonReview, copied: bool) -> list[tuple[str, str, str]]:
    """Per source-file audit: (source file name, action, output file name)."""
    bundle = review.bundle
    if bundle is None:
        return []
    manifest = review.manifest
    actions: list[tuple[str, str, str]] = []
    typed = (
        ("Front", bundle.front, manifest.front_output_name),
        ("Last", bundle.back, manifest.back_output_name),
        ("Blank", bundle.blank, manifest.blank_output_name),
        ("Photo", bundle.photo, manifest.photo_output_name),
    )
    for label, sources, base in typed:
        multiple = len(sources) > 1
        for index, source in enumerate(sources, start=1):
            ext = source.path.suffix.lower()
            if ext in (".jpg", ".jpeg"):
                action, out_ext = "kept (JPG)", ext
            elif ext == ".png":
                action, out_ext = "converted PNG -> JPG", ".jpg"
            elif ext == ".pdf":
                action, out_ext = "extracted PDF -> JPG", ".jpg"
            else:
                action, out_ext = "copied as-is", ext
            suffix = f"_{index:02d}" if multiple else ""
            out = f"{base}{suffix}{out_ext}" if copied else "(not copied - unmatched)"
            actions.append((source.path.name, f"{label} page: {action}", out))
    for path in bundle.ineligible:
        out = "(copied, original format kept)" if copied else "(not copied - unmatched)"
        actions.append((path.name, "copied as-is (format not eligible - not JPG/PDF)", out))
    for path in bundle.excluded:
        actions.append((path.name, "IGNORED (non-passport: voter ID / visa / etc.)", "-"))
    return actions


def _append_file_actions(path: Path, folder_name: str, actions: list[tuple[str, str, str]]) -> None:
    if not actions:
        return
    is_new = not path.exists()
    with path.open("a", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        if is_new:
            writer.writerow(["Folder", "Source File", "Action", "Output File"])
        for source, action, output in actions:
            writer.writerow([folder_name, source, action, output])


def _manifest_from_row(row: SheetRow, folder: Path) -> ManifestRow:
    stem = folder.name.strip() or f"row_{row.row_number}"
    return ManifestRow(
        row_number=row.row_number,
        booking_id=row.booking_id,
        tsf_id=row.tsf_id,
        gender="",
        name=row.name or folder.name,
        expected_passport_no=row.passport_front,
        group="",
        assigned_to="",
        folder_name=row.folder_name or folder.name,
        front_output_name=row.file_front or f"{stem}_passport_front",
        back_output_name=row.file_last or f"{stem}_passport_last",
        blank_output_name=row.file_blank or f"{stem}_passport_blank",
        photo_output_name=row.file_photo or f"{stem}_photo",
        original={},
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Review person-wise passport/photo folders offline and fill the review "
            "spreadsheet (Sheet1). The original workbook is copied, not modified."
        )
    )
    parser.add_argument("--workbook", type=Path, required=True, help="Path to the review .xlsx workbook.")
    parser.add_argument("--folders-root", type=Path, required=True, help="Root whose subfolders are people.")
    parser.add_argument("--output-root", type=Path, required=True, help="Directory for the filled workbook, reports and renamed copies.")
    parser.add_argument("--sheet", default="Sheet1", help="Worksheet to fill (default: Sheet1).")
    parser.add_argument("--config", type=Path, help="Optional JSON configuration file.")
    parser.add_argument("--no-copy", action="store_true", help="Only fill the sheet; do not create renamed copies in the output.")
    parser.add_argument("--ocr-engine", choices=("auto", "paddle", "tesseract"), help="Local OCR engine.")
    parser.add_argument("--skip-local-vision", action="store_true", help="Skip MediaPipe face/hand models.")
    parser.add_argument("--allow-network", action="store_true", help="Do not install the network block. Not recommended for real PII.")
    parser.add_argument("--show-pii-in-console", action="store_true", help="Print folder names while processing.")
    parser.add_argument("--limit", type=int, help="Process only the first N person folders, alphabetically.")
    parser.add_argument("--start-folder", type=int, default=1, help="Start at the Nth folder, alphabetically (1-based).")
    parser.add_argument("--resume", action="store_true", help="Continue a previous run: skip already-processed folders and any folder that crashed.")
    return parser


def _read_lines(path: Path) -> list[str]:
    try:
        return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    except OSError:
        return []


def _append_line(path: Path, text: str) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(text + "\n")


def _setup_engines(config: ReviewConfig, args) -> tuple[MemoizingOCREngine, object | None, bool]:
    offline_strict = not args.allow_network
    if offline_strict:
        enable_strict_offline_mode()
        print("Strict offline mode enabled: outbound network connections are blocked for this process.")
    else:
        print("WARNING: network blocking is disabled. Do not process real PII unless explicitly approved.")

    base_ocr_engine, ocr_warnings = create_ocr_engine(config, args.ocr_engine)
    ocr_engine = MemoizingOCREngine(base_ocr_engine)
    for warning in ocr_warnings:
        print(f"WARNING: {warning}")
    print(f"OCR engine: {ocr_engine.name}")

    visual_analyzer = None
    if not args.skip_local_vision:
        visual_analyzer, visual_warnings = create_visual_analyzer(config)
        for warning in visual_warnings:
            print(f"WARNING: {warning}")
        if visual_analyzer is not None:
            print("Local MediaPipe face/hand models loaded.")
    return ocr_engine, visual_analyzer, offline_strict


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        print("openpyxl is required for spreadsheet mode. Install it with: pip install openpyxl", file=sys.stderr)
        return 2
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    try:
        config = ReviewConfig.from_json(args.config)
    except Exception as exc:
        print(f"Configuration error: {exc}", file=sys.stderr)
        return 2

    # In spreadsheet mode the three-word outcome scheme has no room for a generic
    # "manual review" verdict, so subjective/uncertain items become comments
    # instead of dragging every page and photo to an error.
    config.manual_subjective_photo_checks = False
    config.manual_object_check_when_no_hand_detected = False
    # The workbook stays on the local machine, so passport numbers are shown in full
    # in the comments rather than masked.
    config.mask_passport_numbers = False

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    args.output_root.mkdir(parents=True, exist_ok=True)
    reviewed_path = args.output_root / f"{workbook_path.stem}_reviewed.xlsx"
    documents_root = args.output_root / "reviewed-documents"
    results_csv = args.output_root / "review_results.csv"
    file_actions_csv = args.output_root / "file_actions.csv"
    unmatched_folders_file = args.output_root / "_unmatched_folders.tsv"
    ckpt_done = args.output_root / "_progress_done.txt"
    ckpt_current = args.output_root / "_progress_current.txt"
    ckpt_skip = args.output_root / "_progress_skip.txt"

    # Resume bookkeeping. A folder is added to the skip list if it crashed the
    # process last time (its name is left in _progress_current.txt because a
    # segmentation fault in a native library cannot be caught in Python).
    processed: set[str] = set(_read_lines(ckpt_done)) if args.resume else set()
    skip: set[str] = set(_read_lines(ckpt_skip)) if args.resume else set()
    if args.resume:
        crasher = ckpt_current.read_text(encoding="utf-8").strip() if ckpt_current.exists() else ""
        if crasher and crasher not in processed and crasher not in skip:
            skip.add(crasher)
            _append_line(ckpt_skip, crasher)
            print("Auto-skipping one folder that crashed the previous run.")
    else:
        for stale in (ckpt_done, ckpt_current, ckpt_skip, results_csv, file_actions_csv, unmatched_folders_file):
            try:
                stale.unlink()
            except FileNotFoundError:
                pass

    # Read computed values (formulas) from the original; write into the in-progress
    # reviewed copy when resuming so fills accumulate across runs.
    write_source = reviewed_path if (args.resume and reviewed_path.exists()) else workbook_path
    try:
        values_wb = openpyxl.load_workbook(workbook_path, data_only=True)
        workbook = openpyxl.load_workbook(write_source)
    except Exception as exc:
        print(f"Could not open workbook: {exc}", file=sys.stderr)
        return 2
    if args.sheet not in workbook.sheetnames:
        print(f"Worksheet {args.sheet!r} not found. Available: {workbook.sheetnames}", file=sys.stderr)
        return 2
    values_ws = values_wb[args.sheet]
    worksheet = workbook[args.sheet]

    header_row = _find_header_row(values_ws)
    cols = _resolve_columns(values_ws, header_row)
    front_review_col = _front_review_column(worksheet, header_row)
    email_col = _email_column(worksheet, header_row)
    sheet_rows = _load_sheet_rows(values_ws, header_row, cols)
    index = MatchIndex.build(sheet_rows)
    review_cols = (front_review_col, cols.passport_back, cols.passport_blank, cols.photo_dimension)
    for row in sheet_rows:
        if any(worksheet.cell(row=row.row_number, column=c).value for c in review_cols):
            row.filled = True  # filled by a previous run
    print(f"Spreadsheet rows loaded: {len(sheet_rows)} (header row {header_row}).")
    if args.resume:
        print(f"Resuming: {len(processed)} folder(s) already done, {len(skip)} skipped.")

    try:
        folders = list_person_folders(args.folders_root, args.output_root)
    except Exception as exc:
        print(f"Folder error: {exc}", file=sys.stderr)
        return 2

    start = max(args.start_folder - 1, 0)
    folders = folders[start:]
    if args.limit is not None:
        folders = folders[: max(args.limit, 0)]
    if not folders:
        print("No person folders were found.", file=sys.stderr)
        return 2
    print(f"Person folders discovered: {len(folders)}")

    copy_enabled = not args.no_copy
    ocr_engine, visual_analyzer, offline_strict = _setup_engines(config, args)

    processed_this_run = 0
    save_failed = False

    try:
        for position, folder in enumerate(folders, start=1):
            if folder.name in processed or folder.name in skip:
                continue
            label = folder.name if args.show_pii_in_console else f"person folder {position}"
            print(f"[{position}/{len(folders)}] Reviewing {label}...")
            # Record the folder before touching native code; if it segfaults, the
            # next --resume run finds this name and skips it.
            ckpt_current.write_text(folder.name, encoding="utf-8")
            try:
                tsf_ids, booking_ids = _ids_from_folder(folder)
                tentative = _single([r for tsf in tsf_ids for r in index.by_tsf.get(tsf, [])])
                if tentative is None:
                    tentative = _booking_row(index, booking_ids)

                manifest = _manifest_from_row(tentative, folder) if tentative else None
                review = review_person_folder(
                    folder=folder,
                    row_number=start + position,
                    output_root=documents_root,
                    config=config,
                    ocr_engine=ocr_engine,
                    visual_analyzer=visual_analyzer,
                    copy_renamed=False,
                    manifest=manifest,
                    rename_in_place=False,
                )

                decision = _decide_row(
                    review.detected_passport_number, tentative, tsf_ids, booking_ids, folder, index, sheet_rows
                )
                if decision.row is None:
                    print(f"    {review.overall.value} (no spreadsheet row matched)")
                    _append_line(unmatched_folders_file, f"{folder.name}\t{review.detected_passport_number}")
                else:
                    row = decision.row
                    manifest = _manifest_from_row(row, folder)
                    review.manifest = manifest
                    note = decision.note
                    if copy_enabled and review.bundle is not None:
                        try:
                            review.output_files = export_bundle(review.bundle, manifest, documents_root, config)
                        except Exception as exc:
                            note = (note + f" Could not copy output files: {exc}").strip()
                    bundle = review.bundle
                    fronts = bundle.front if bundle else []
                    backs = bundle.back if bundle else []
                    blanks = bundle.blank if bundle else []
                    photos = bundle.photo if bundle else []
                    worksheet.cell(row=row.row_number, column=front_review_col).value = _document_cell(review.front.status, fronts)
                    worksheet.cell(row=row.row_number, column=cols.passport_back).value = _document_cell(review.back.status, backs)
                    worksheet.cell(row=row.row_number, column=cols.passport_blank).value = _document_cell(review.blank.status, blanks)
                    worksheet.cell(row=row.row_number, column=cols.photo_dimension).value = _document_cell(review.photo.status, photos)
                    worksheet.cell(row=row.row_number, column=cols.comments).value = _sheet_comments(review, note)
                    email = _email_text(row.name, review, config)
                    if email:
                        email_cell = worksheet.cell(row=row.row_number, column=email_col)
                        email_cell.value = email
                        email_cell.alignment = wrap_alignment
                    row.filled = True
                    print(f"    {review.overall.value} -> row {row.row_number} (matched by {decision.method})")

                # Persist the workbook BEFORE marking the folder done, so a later
                # crash can never lose a fill that is recorded as complete.
                try:
                    workbook.save(reviewed_path)
                except Exception as exc:
                    print(f"Could not save workbook (is it open in Excel?): {exc}", file=sys.stderr)
                    save_failed = True
                    break
                append_result_row(results_csv, review)
                _append_file_actions(
                    file_actions_csv, folder.name,
                    _file_actions(review, copied=(decision.row is not None and copy_enabled)),
                )
                _append_line(ckpt_done, folder.name)
                processed.add(folder.name)
                processed_this_run += 1
            except KeyboardInterrupt:
                print("Interrupted by user.", file=sys.stderr)
                break
            except Exception as exc:
                # Recoverable (non-crash) error: skip this folder and keep going.
                print(f"    Unexpected error: {exc}", file=sys.stderr)
                _append_line(ckpt_skip, folder.name)
                skip.add(folder.name)
    finally:
        if visual_analyzer is not None:
            visual_analyzer.close()

    if not save_failed:
        try:
            ckpt_current.unlink()  # clean finish: nothing in progress
        except FileNotFoundError:
            pass

    if save_failed:
        print("Stopped: the workbook could not be saved. Close it in Excel, then re-run with --resume.", file=sys.stderr)
        return 1

    remaining = [f for f in folders if f.name not in processed and f.name not in skip]
    filled_total = sum(1 for row in sheet_rows if row.filled)

    # Final reports are built from the complete, appended results CSV and the final
    # workbook, so they are correct even across several resumed runs.
    _finalize_reports(results_csv, args.output_root, unmatched_folders_file, sheet_rows,
                      ocr_engine.name, offline_strict, filled_total)

    print(f"\nProcessed this run: {processed_this_run}. Remaining: {len(remaining)}.")
    print(f"Rows filled (total): {filled_total} / {len(sheet_rows)}")
    print(f"Filled workbook: {reviewed_path}")
    if copy_enabled:
        print(f"Renamed copies:  {documents_root}")
    print(f"Full report:     {results_csv}")
    print(f"File actions:    {file_actions_csv}")
    print(f"Manual queue:    {args.output_root / 'manual_review_queue.csv'}")
    print(f"Unmatched list:  {args.output_root / 'unmatched.csv'}")
    if remaining:
        print(f"\nNOTE: {len(remaining)} folder(s) still to process. Re-run the same command with --resume.")
        return 3
    if skip:
        print(f"NOTE: {len(skip)} folder(s) were skipped (crashed or errored) - listed in {ckpt_skip.name}.")
    return 0


def _finalize_reports(results_csv: Path, output_root: Path, unmatched_folders_file: Path,
                      sheet_rows: list[SheetRow], ocr_name: str, offline_strict: bool, filled_total: int) -> None:
    from collections import Counter

    counts: Counter = Counter()
    manual: list[dict] = []
    if results_csv.exists():
        with results_csv.open(encoding="utf-8-sig", newline="") as handle:
            for record in csv.DictReader(handle):
                overall = record.get("Overall Review", "")
                counts[overall] += 1
                if overall and overall != "ok":
                    manual.append(record)

    with (output_root / "review_summary.txt").open("w", encoding="utf-8") as handle:
        handle.write("Passport Review Summary\n=======================\n\n")
        handle.write(f"OCR engine: {ocr_name}\n")
        handle.write(f"Strict offline network block: {'enabled' if offline_strict else 'disabled'}\n")
        handle.write(f"Rows filled: {filled_total} / {len(sheet_rows)}\n\n")
        for status, count in sorted(counts.items()):
            handle.write(f"{status}: {count}\n")

    with (output_root / "manual_review_queue.csv").open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Person Folder", "Overall Review", "Front", "Back", "Blank Pages", "Photo", "Review Comments"])
        for record in manual:
            writer.writerow([
                record.get("Person Folder", ""), record.get("Overall Review", ""),
                record.get("Passport Front Review", ""), record.get("Passport Last Page Review", ""),
                record.get("Passport Blank Pages Review", ""), record.get("Photo Full Review", ""),
                record.get("Review Comments", ""),
            ])

    with (output_root / "unmatched.csv").open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Type", "Folder / Row", "Passport Number", "Note"])
        for line in _read_lines(unmatched_folders_file):
            name, _, number = line.partition("\t")
            writer.writerow(["folder without row", name, number, "No spreadsheet row matched this folder."])
        for row in sheet_rows:
            if not row.filled:
                writer.writerow(["row without folder", row.folder_name or row.name or f"row {row.row_number}", "", "No reviewed folder was matched to this row."])
